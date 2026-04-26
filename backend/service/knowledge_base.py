from __future__ import annotations

import csv
import hashlib
import json
import logging
import math
import re
from dataclasses import asdict, dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

import jieba
from rank_bm25 import BM25Okapi

from config import get_settings
from graph.llm import build_embedding_config_from_settings, get_embedding_model

logger = logging.getLogger(__name__)

_WORD_RE = re.compile(r"[A-Za-z0-9_]+")
_HAS_CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
_WHITESPACE_RE = re.compile(r"\s+")
_SUPPORTED_SUFFIXES = {".md", ".txt", ".json", ".csv", ".pdf", ".xlsx"}


@dataclass
class KnowledgeChunk:
    chunk_id: str
    source: str
    title: str
    text: str


def _safe_json_dump(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _normalize_text(text: str) -> str:
    return _WHITESPACE_RE.sub(" ", text or "").strip()


def _tokenize(text: str) -> list[str]:
    content = (text or "").strip().lower()
    if not content:
        return []

    tokens: list[str] = []
    if _HAS_CJK_RE.search(content):
        tokens.extend(token.strip() for token in jieba.lcut(content) if token.strip())
    tokens.extend(token.strip() for token in _WORD_RE.findall(content) if token.strip())
    return tokens


class KnowledgeBase:
    def __init__(self, root_dir: Path) -> None:
        self.root_dir = root_dir.resolve()
        self.knowledge_dir = self.root_dir / "knowledge"
        self.storage_dir = self.root_dir / "storage" / "knowledge_rag"
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        self._chunks: list[KnowledgeChunk] = []
        self._embeddings: list[list[float]] = []
        self._bm25: BM25Okapi | None = None
        self._digest: str = ""

    @property
    def _index_path(self) -> Path:
        return self.storage_dir / "index.json"

    def _settings(self):
        return get_settings()

    def _supports_dense(self) -> bool:
        settings = self._settings()
        dense_enabled = bool(settings.knowledge_dense_enabled)
        return dense_enabled and bool(settings.embedding_api_key)

    def _fingerprint(self) -> str:
        settings = self._settings()
        payload: list[str] = [
            str(settings.knowledge_chunk_size),
            str(settings.knowledge_chunk_overlap),
            str(settings.knowledge_max_chunks_per_file),
            str(bool(settings.knowledge_dense_enabled)),
        ]
        for path in sorted(self.knowledge_dir.rglob("*")):
            if path.is_file():
                stat = path.stat()
                payload.append(
                    f"{path.relative_to(self.root_dir)}:{stat.st_mtime_ns}:{stat.st_size}"
                )
        return hashlib.md5("\n".join(payload).encode("utf-8")).hexdigest()

    def _embed_documents(self, texts: list[str]) -> list[list[float]]:
        if not texts or not self._supports_dense():
            return []

        embedding_model = get_embedding_model(
            build_embedding_config_from_settings(self._settings())
        )
        batch_size = 32
        results: list[list[float]] = []
        for start in range(0, len(texts), batch_size):
            batch = texts[start : start + batch_size]
            results.extend(embedding_model.embed_documents(batch))
        return results

    def _embed_query(self, query: str) -> list[float] | None:
        if not query.strip() or not self._supports_dense():
            return None

        embedding_model = get_embedding_model(
            build_embedding_config_from_settings(self._settings())
        )
        return embedding_model.embed_query(query)

    def _read_text_file(self, path: Path) -> str:
        return path.read_text(encoding="utf-8", errors="ignore")

    def _read_json_file(self, path: Path) -> str:
        try:
            payload = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
            return _safe_json_dump(payload)
        except json.JSONDecodeError:
            return path.read_text(encoding="utf-8", errors="ignore")

    def _read_csv_file(self, path: Path) -> str:
        rows: list[str] = []
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader, start=1):
                if not row:
                    continue
                rows.append(f"row {index}: " + " | ".join(cell.strip() for cell in row))
        return "\n".join(rows)

    def _read_pdf_file(self, path: Path) -> str:
        try:
            from pypdf import PdfReader
        except ImportError:
            logger.warning("Skipping PDF %s because pypdf is not installed", path.name)
            return ""

        try:
            reader = PdfReader(str(path))
        except Exception as exc:  # pragma: no cover - depends on file contents
            logger.warning("Failed to open PDF %s: %s", path.name, exc)
            return ""

        pages: list[str] = []
        for page_index, page in enumerate(reader.pages, start=1):
            try:
                text = page.extract_text() or ""
            except Exception:
                text = ""
            text = _normalize_text(text)
            if text:
                pages.append(f"[page {page_index}] {text}")
        return "\n".join(pages)

    def _read_xlsx_file(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("Skipping spreadsheet %s because openpyxl is not installed", path.name)
            return ""

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - depends on file contents
            logger.warning("Failed to open spreadsheet %s: %s", path.name, exc)
            return ""

        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# Sheet: {sheet.title}")
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(cell).strip() for cell in row if cell not in (None, "")]
                if not values:
                    continue
                lines.append(f"row {row_index}: " + " | ".join(values))
        return "\n".join(lines)

    def _read_document(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in {".md", ".txt"}:
            return self._read_text_file(path)
        if suffix == ".json":
            return self._read_json_file(path)
        if suffix == ".csv":
            return self._read_csv_file(path)
        if suffix == ".pdf":
            return self._read_pdf_file(path)
        if suffix == ".xlsx":
            return self._read_xlsx_file(path)
        return ""

    def _chunk_document(self, path: Path, text: str) -> list[KnowledgeChunk]:
        settings = self._settings()
        chunk_size = max(200, int(settings.knowledge_chunk_size))
        overlap = max(0, min(chunk_size // 2, int(settings.knowledge_chunk_overlap)))
        max_chunks = max(1, int(settings.knowledge_max_chunks_per_file))

        cleaned = _normalize_text(text)
        if not cleaned:
            return []

        relative_path = str(path.relative_to(self.root_dir)).replace("\\", "/")
        step = max(50, chunk_size - overlap)
        chunks: list[KnowledgeChunk] = []
        for index, start in enumerate(range(0, len(cleaned), step), start=1):
            snippet = cleaned[start : start + chunk_size].strip()
            if not snippet:
                continue
            chunk_id = hashlib.md5(
                f"{relative_path}:{index}:{snippet[:120]}".encode("utf-8")
            ).hexdigest()
            chunks.append(
                KnowledgeChunk(
                    chunk_id=chunk_id,
                    source=relative_path,
                    title=path.name,
                    text=snippet,
                )
            )
            if len(chunks) >= max_chunks:
                break
        return chunks

    def _load_documents(self) -> list[KnowledgeChunk]:
        chunks: list[KnowledgeChunk] = []
        self.knowledge_dir.mkdir(parents=True, exist_ok=True)

        for path in sorted(self.knowledge_dir.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in _SUPPORTED_SUFFIXES:
                continue
            text = self._read_document(path)
            if not text:
                continue
            chunks.extend(self._chunk_document(path, text))
        return chunks

    def _persist_index(self) -> None:
        payload = {
            "digest": self._digest,
            "chunks": [asdict(chunk) for chunk in self._chunks],
            "embeddings": self._embeddings,
        }
        self._index_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _load_cached_index(self, digest: str) -> bool:
        if not self._index_path.exists():
            return False

        try:
            payload = json.loads(self._index_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return False

        if payload.get("digest") != digest:
            return False

        self._digest = digest
        self._chunks = [KnowledgeChunk(**item) for item in payload.get("chunks", [])]
        self._embeddings = payload.get("embeddings", []) or []
        self._build_bm25()
        return True

    def _build_bm25(self) -> None:
        corpus = [_tokenize(chunk.text) for chunk in self._chunks]
        if corpus and any(tokens for tokens in corpus):
            self._bm25 = BM25Okapi(corpus)
        else:
            self._bm25 = None

    def rebuild(self) -> None:
        self.knowledge_dir.mkdir(parents=True, exist_ok=True)
        self._digest = self._fingerprint()
        self._chunks = self._load_documents()
        self._build_bm25()
        self._embeddings = []

        if self._chunks and self._supports_dense():
            try:
                self._embeddings = self._embed_documents([chunk.text for chunk in self._chunks])
            except Exception as exc:
                logger.warning("Dense knowledge index build failed, fallback to keyword only: %s", exc)
                self._embeddings = []

        self._persist_index()

    def ensure_loaded(self) -> None:
        digest = self._fingerprint()
        if digest == self._digest and self._chunks:
            return
        if self._load_cached_index(digest):
            return
        self.rebuild()

    def _keyword_scores(self, query: str) -> list[float]:
        query_tokens = _tokenize(query)
        if not query_tokens:
            return [0.0 for _ in self._chunks]

        lexical_scores: list[float] = []
        for chunk in self._chunks:
            chunk_tokens = set(_tokenize(chunk.text))
            overlap = len(chunk_tokens.intersection(query_tokens))
            lexical_scores.append(overlap / max(1, len(set(query_tokens))))

        if self._bm25 is None:
            return lexical_scores

        bm25_scores = [float(score) for score in self._bm25.get_scores(query_tokens)]
        if max(bm25_scores, default=0.0) <= 0:
            return lexical_scores

        return [
            bm25_scores[index] + 0.15 * lexical_scores[index]
            for index in range(len(self._chunks))
        ]

    def _dense_scores(self, query: str) -> list[float]:
        if not self._embeddings:
            return [0.0 for _ in self._chunks]

        try:
            query_embedding = self._embed_query(query)
        except Exception as exc:
            logger.warning("Knowledge query embedding failed, fallback to keyword only: %s", exc)
            return [0.0 for _ in self._chunks]

        if not query_embedding:
            return [0.0 for _ in self._chunks]

        def cosine(left: list[float], right: list[float]) -> float:
            if not left or not right:
                return 0.0
            dot = sum(lv * rv for lv, rv in zip(left, right))
            left_norm = math.sqrt(sum(lv * lv for lv in left))
            right_norm = math.sqrt(sum(rv * rv for rv in right))
            if left_norm == 0 or right_norm == 0:
                return 0.0
            return dot / (left_norm * right_norm)

        return [cosine(query_embedding, embedding) for embedding in self._embeddings]

    @staticmethod
    def _normalize_scores(scores: list[float]) -> list[float]:
        if not scores:
            return []
        highest = max(scores)
        lowest = min(scores)
        if highest == lowest:
            return [1.0 if highest > 0 else 0.0 for _ in scores]
        return [(score - lowest) / (highest - lowest) for score in scores]

    def search(self, query: str, top_k: int | None = None) -> list[dict[str, Any]]:
        self.ensure_loaded()
        if not query.strip() or not self._chunks:
            return []

        settings = self._settings()
        final_top_k = int(top_k or settings.knowledge_top_k)
        dense_weight = 0.45 if self._embeddings else 0.0
        keyword_weight = 1.0 - dense_weight

        keyword_scores = self._normalize_scores(self._keyword_scores(query))
        dense_scores = self._normalize_scores(self._dense_scores(query))

        ranked: list[dict[str, Any]] = []
        for index, chunk in enumerate(self._chunks):
            score = keyword_weight * keyword_scores[index]
            if dense_scores:
                score += dense_weight * dense_scores[index]
            if score <= 0:
                continue
            ranked.append(
                {
                    "text": chunk.text,
                    "score": round(score, 6),
                    "source": chunk.source,
                    "kind": "knowledge",
                }
            )

        ranked.sort(key=lambda item: item["score"], reverse=True)
        return ranked[: max(1, final_top_k)]

    def format_context(self, results: list[dict[str, Any]]) -> str:
        if not results:
            return ""

        lines = ["[Knowledge base context]"]
        for index, item in enumerate(results, start=1):
            lines.append(
                f"{index}. Source: {item['source']}\n{str(item['text']).strip()}"
            )
        return "\n\n".join(lines)


@lru_cache(maxsize=4)
def get_knowledge_base(root_dir: Path) -> KnowledgeBase:
    return KnowledgeBase(root_dir)
